import streamlit as st
import openpyxl
from fpdf import FPDF
import pandas as pd
import openpyxl.utils # Para get_column_letter
import os
import io

# --- Helper Functions (Copied and adapted from your script) ---
COLOR_TO_RISK_LEVEL = {
    "FF00FF00": "ACEPTABLE", "00FF00": "ACEPTABLE", "FF008000": "ACEPTABLE", "FF92D050": "ACEPTABLE",
    "FFFFFF00": "INTERMEDIO", "FFFFC000": "INTERMEDIO", "FFFFA500": "INTERMEDIO",
    "FFFF0000": "CRÍTICO",
}
user_logs = []

def log_to_streamlit_and_console(message, level="info"):
    formatted_message = f"[{level.upper()}] {message}"
    user_logs.append(formatted_message)
    print(formatted_message)

def add_table_to_pdf(pdf, headers, data_rows, col_widths_list=None):
    if not data_rows and not headers:
        return

    line_height_header = 8 # Aumentado de 7
    line_height_data = 7   # Aumentado de 6
    
    pdf.set_font("Arial", "B", 8) # Aumentado de 7 para cabeceras
    page_width = pdf.w - 2 * pdf.l_margin
    num_cols = len(headers) if headers else (len(data_rows[0]) if data_rows else 0)
    if num_cols == 0: return

    actual_col_widths = col_widths_list
    if col_widths_list is None or len(col_widths_list) != num_cols:
        default_col_width = page_width / num_cols
        actual_col_widths = [default_col_width] * num_cols
        if col_widths_list:
            log_to_streamlit_and_console(f"Anchos de columna no coinciden ({len(col_widths_list)} vs {num_cols}). Usando por defecto.", "warning")

    if headers:
        current_x_start_of_headers = pdf.l_margin
        current_y_start_of_headers = pdf.get_y()
        max_y_after_header_multicell = current_y_start_of_headers
        offset_x = 0
        for i, header_text in enumerate(headers):
            pdf.set_xy(current_x_start_of_headers + offset_x, current_y_start_of_headers)
            pdf.multi_cell(actual_col_widths[i], line_height_header, str(header_text), border=1, align="C")
            max_y_after_header_multicell = max(max_y_after_header_multicell, pdf.get_y())
            offset_x += actual_col_widths[i]
        pdf.set_xy(pdf.l_margin, max_y_after_header_multicell)

    pdf.set_font("Arial", "", 7) # Aumentado de 6 para datos
    for row in data_rows:
        if pdf.get_y() + line_height_data > pdf.page_break_trigger:
            pdf.add_page()
            if headers:
                pdf.set_font("Arial", "B", 8) # Re-aplicar fuente de cabecera
                h_current_x_start = pdf.l_margin
                h_current_y_start = pdf.get_y()
                h_max_y_after_multicell = h_current_y_start
                h_offset_x = 0
                for i, header_text in enumerate(headers):
                    pdf.set_xy(h_current_x_start + h_offset_x, h_current_y_start)
                    pdf.multi_cell(actual_col_widths[i], line_height_header, str(header_text), border=1, align="C")
                    h_max_y_after_multicell = max(h_max_y_after_multicell, pdf.get_y())
                    h_offset_x += actual_col_widths[i]
                pdf.set_xy(pdf.l_margin, h_max_y_after_multicell)
                pdf.set_font("Arial", "", 7) # Re-aplicar fuente de datos
        
        current_x_start_of_row = pdf.get_x()
        current_y_start_of_row = pdf.get_y()
        max_y_after_multicell = current_y_start_of_row
        offset_x_data = 0
        for i, cell_text in enumerate(row):
            pdf.set_xy(current_x_start_of_row + offset_x_data, current_y_start_of_row)
            pdf.multi_cell(actual_col_widths[i], line_height_data, str(cell_text), border=1, align="L")
            max_y_after_multicell = max(max_y_after_multicell, pdf.get_y())
            offset_x_data += actual_col_widths[i]
        pdf.set_xy(pdf.l_margin, max_y_after_multicell)

def add_structured_entry_to_pdf(pdf, entry_data_list, column_groups_with_headers, entry_id_text=None):
    if not entry_data_list or not column_groups_with_headers: return

    if entry_id_text:
        pdf.set_font("Arial", "B", 10) # Aumentado de 9
        pdf.cell(0, 8, txt=entry_id_text, ln=True, align="L") # Aumentado de 7
        pdf.ln(1)

    line_height = 6 # Aumentado de 5
    label_width_ratio = 0.30
    value_width_ratio = 0.68
    page_width = pdf.w - 2 * pdf.l_margin
    label_col_width = page_width * label_width_ratio
    value_col_width = page_width * value_width_ratio

    for group_idx, group in enumerate(column_groups_with_headers):
        if not group: continue
        if group_idx > 0: pdf.ln(line_height / 2)

        for header_name, col_idx in group:
            value = str(entry_data_list[col_idx]) if col_idx < len(entry_data_list) else ""
            y_start_pair = pdf.get_y()
            if y_start_pair + line_height > pdf.page_break_trigger and pdf.auto_page_break:
                pdf.add_page()
                y_start_pair = pdf.get_y()

            pdf.set_font("Arial", "B", 8) # Aumentado de 7
            pdf.set_x(pdf.l_margin)
            y_before_label = pdf.get_y()
            pdf.multi_cell(label_col_width, line_height, txt=f"{header_name}:", border=0, align="L")
            y_after_label = pdf.get_y()
            
            pdf.set_xy(pdf.l_margin + label_col_width, y_before_label)
            pdf.set_font("Arial", "", 8) # Aumentado de 7
            pdf.multi_cell(value_col_width, line_height, txt=value, border=0, align="L")
            y_after_value = pdf.get_y()
            pdf.set_y(max(y_after_label, y_after_value))

    pdf.ln(1)
    current_y_before_line = pdf.get_y()
    if current_y_before_line + 2 < pdf.page_break_trigger: # +2 para línea y espacio
        pdf.line(pdf.l_margin, current_y_before_line, pdf.w - pdf.r_margin, current_y_before_line)
        pdf.ln(2)
    else:
        pdf.ln(1)

def add_resumen_page_to_pdf(pdf, resumen_data, agentes_ordenados):
    if not resumen_data:
        log_to_streamlit_and_console("No hay datos para generar la página de resumen.", "info")
        return

    # No se añade página aquí, se asume que ya se hizo antes de llamar esta función
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, txt="Resumen de Niveles de Riesgo por Puesto de Trabajo", ln=True, align="C")
    pdf.ln(5)

    headers_resumen = ["N°", "Área", "Puesto", "Tarea"] + agentes_ordenados
    col_widths_resumen = [15, 60, 60, 80] + [25] * len(agentes_ordenados)
    data_rows_resumen = []
    
    # Ordenar por el N° (primer elemento de la tupla clave)
    # Asegurarse de que la clave es una tupla y el primer elemento es el N°
    def get_sort_key(item):
        key_tuple = item[0]
        if isinstance(key_tuple, tuple) and len(key_tuple) > 0:
            num_str = str(key_tuple[0])
            if num_str.replace('.', '', 1).isdigit():
                return int(num_str.split('.')[0])
        return float('inf') # Poner al final si no es un número válido

    sorted_casos = sorted(resumen_data.items(), key=get_sort_key)

    for caso_key, niveles_riesgo in sorted_casos:
        fila_pdf = list(caso_key) + [niveles_riesgo.get(agente, "AUSENTE") for agente in agentes_ordenados]
        data_rows_resumen.append(fila_pdf)

    if data_rows_resumen:
        add_table_to_pdf(pdf, headers_resumen, data_rows_resumen, col_widths_resumen)
    else:
        pdf.set_font("Arial", "", 10)
        pdf.cell(0,10, txt="No se encontraron casos para resumir.", ln=True, align="L")

# --- Funciones de recolección de datos ---
def collect_hoja1_data(wb, mapeo_pdf_hoja1):
    collected_data = []
    try:
        hoja1_openpyxl = wb["1"]
        log_to_streamlit_and_console("Recolectando datos de Hoja 1...", "info")
        for seccion_titulo, campos in mapeo_pdf_hoja1.items():
            texto_seccion_completo = ""
            for etiqueta, (fila_excel, col_excel_char) in campos.items():
                valor_crudo_h1 = hoja1_openpyxl[f"{col_excel_char}{fila_excel}"].value
                valor_str_h1 = ""
                if valor_crudo_h1 is not None:
                    valor_str_temp_h1 = str(valor_crudo_h1).strip()
                    if valor_str_temp_h1 != "0": valor_str_h1 = valor_str_temp_h1
                if valor_str_h1: texto_seccion_completo += f"{etiqueta}: {valor_str_h1}  "
            collected_data.append({"titulo": seccion_titulo, "texto": texto_seccion_completo.strip()})
    except KeyError:
        log_to_streamlit_and_console("No se encontró la Hoja '1'. Se omitirá su recolección.", "warning")
    except Exception as e:
        log_to_streamlit_and_console(f"Error recolectando datos de Hoja '1': {e}", "error")
    return collected_data

def collect_hoja2_data(wb, hoja2_col_groups_structured, resumen_riesgos_data_ref, agentes_riesgo_ordenados):
    collected_data = []
    all_apt_cases_from_hoja2 = set() # Para asegurar que el resumen tenga todos los casos de Hoja 2
    COL_AREA, COL_PUESTO, COL_TAREA = 3, 4, 5
    try:
        hoja2 = wb["2"]
        log_to_streamlit_and_console("Recolectando datos de Hoja 2...", "info")
        for fila_idx in range(13, 114):
            val_a_obj = hoja2.cell(row=fila_idx, column=COL_AREA).value
            val_p_obj = hoja2.cell(row=fila_idx, column=COL_PUESTO).value
            val_t_obj = hoja2.cell(row=fila_idx, column=COL_TAREA).value
            val_a_str = str(val_a_obj).strip() if val_a_obj is not None else ""
            val_p_str = str(val_p_obj).strip() if val_p_obj is not None else ""
            val_t_str = str(val_t_obj).strip() if val_t_obj is not None else ""

            if not (val_a_str and val_a_str != "0" and val_p_str and val_p_str != "0" and val_t_str and val_t_str != "0"):
                continue
            
            current_row_values = [str(hoja2.cell(row=fila_idx, column=col_idx).value or "") for col_idx in range(2, 21)]
            if any(val.strip() for val in current_row_values):
                entry_id_text = f"Registro Puesto (Hoja 2) N°: {current_row_values[0]}" if current_row_values else f"Registro Puesto (Hoja 2) Fila {fila_idx}"
                collected_data.append({"id_text": entry_id_text, "values": current_row_values, "groups": hoja2_col_groups_structured})
                
                # Para el resumen: N°, Área, Puesto, Tarea
                caso_key_hoja2 = (current_row_values[0], val_a_str, val_p_str, val_t_str)
                all_apt_cases_from_hoja2.add(caso_key_hoja2)
                if caso_key_hoja2 not in resumen_riesgos_data_ref:
                     resumen_riesgos_data_ref[caso_key_hoja2] = {agente: "AUSENTE" for agente in agentes_riesgo_ordenados}

    except KeyError:
        log_to_streamlit_and_console("No se encontró la Hoja '2'. Se omitirá su recolección.", "warning")
    except Exception as e:
        log_to_streamlit_and_console(f"Error recolectando datos de Hoja '2': {e}", "error")
    return collected_data, all_apt_cases_from_hoja2


def collect_factor_data(wb, num_hoja, config, resumen_riesgos_data_ref, agentes_riesgo_ordenados):
    collected_data_for_factor = []
    COL_AREA, COL_PUESTO, COL_TAREA = 3, 4, 5
    col_indices = [2, 3, 4, 5, config["r_col"]]
    try:
        hoja_actual = wb[num_hoja]
        log_to_streamlit_and_console(f"Recolectando datos de Hoja {num_hoja} ({config['nombre']})...", "info")
        for fila_idx in range(config["r_filas"][0], config["r_filas"][1]):
            val_a_obj = hoja_actual.cell(row=fila_idx, column=COL_AREA).value
            val_p_obj = hoja_actual.cell(row=fila_idx, column=COL_PUESTO).value
            val_t_obj = hoja_actual.cell(row=fila_idx, column=COL_TAREA).value
            val_a_str = str(val_a_obj).strip() if val_a_obj is not None else ""
            val_p_str = str(val_p_obj).strip() if val_p_obj is not None else ""
            val_t_str = str(val_t_obj).strip() if val_t_obj is not None else ""

            if not (val_a_str and val_a_str != "0" and val_p_str and val_p_str != "0" and val_t_str and val_t_str != "0"):
                continue
            
            current_row_values = []
            risk_level_text = "No Determinado"
            for _, col_idx_excel in enumerate(col_indices):
                celda = hoja_actual.cell(row=fila_idx, column=col_idx_excel)
                if col_idx_excel == config["r_col"]:
                    valor_crudo = celda.value
                    valor_str = str(valor_crudo).strip().lower() if valor_crudo is not None else ""
                    if "no crítico" in valor_str: risk_level_text = "MEDIO"
                    elif "aceptable" in valor_str: risk_level_text = "BAJO"
                    elif "crítico" in valor_str: risk_level_text = "ALTO"
                    elif "intermedio" in valor_str: risk_level_text = "MEDIO"
                    elif valor_str: risk_level_text = str(valor_crudo).strip()
                    current_row_values.append(risk_level_text)
                else:
                    current_row_values.append(str(celda.value or ""))
            
            if all(str(current_row_values[j]).strip() for j in range(4)) and any(str(val).strip() for val in current_row_values):
                collected_data_for_factor.append(current_row_values)
                caso_key_factor = tuple(str(current_row_values[j]) for j in range(4)) # N°, Área, Puesto, Tarea
                
                if caso_key_factor not in resumen_riesgos_data_ref:
                    log_to_streamlit_and_console(f"Caso {caso_key_factor} de Hoja {num_hoja} no estaba en Hoja 2. Agregando al resumen.", "warning")
                    resumen_riesgos_data_ref[caso_key_factor] = {agente: "AUSENTE" for agente in agentes_riesgo_ordenados}
                resumen_riesgos_data_ref[caso_key_factor][agentes_riesgo_ordenados[config["agente_idx"]]] = risk_level_text
        
        if not collected_data_for_factor:
            log_to_streamlit_and_console(f"No se recolectaron datos para Hoja {num_hoja} ({config['nombre']}) después del filtrado.", "info")

    except KeyError:
        log_to_streamlit_and_console(f"No se encontró la Hoja '{num_hoja}'. Se omitirá su recolección.", "warning")
    except Exception as e:
        log_to_streamlit_and_console(f"Error recolectando datos de Hoja '{num_hoja}': {e}", "error")
    return collected_data_for_factor

# --- Funciones de escritura al PDF ---
def write_hoja1_to_pdf(pdf, collected_data_h1):
    if not collected_data_h1: return
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, txt="Datos de Hoja 1", ln=True, align="L")
    pdf.ln(2)
    line_height_h1 = 7 # Aumentado
    for item in collected_data_h1:
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 8, txt=item["titulo"], ln=True, align="L")
        if item["texto"]:
            pdf.set_font("Arial", "", 8) # Aumentado
            pdf.multi_cell(0, line_height_h1, txt=item["texto"], border=0, align="L")
            pdf.ln(line_height_h1 / 2)
        pdf.ln(1)
    pdf.ln(3)

def write_hoja2_to_pdf(pdf, collected_data_h2):
    if not collected_data_h2: return
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, txt="Datos de Hoja 2", ln=True, align="L")
    pdf.set_font("Arial", "", 8) # Reset
    for item in collected_data_h2:
        add_structured_entry_to_pdf(pdf, item["values"], item["groups"], item["id_text"])

def write_factor_data_to_pdf(pdf, num_hoja, config, data_for_this_factor):
    if not data_for_this_factor: return
    headers = ["N°", "Área", "Puesto", "Tarea", "Nivel de Riesgo"]
    col_widths = [(pdf.w - 2 * pdf.l_margin) / len(headers)] * len(headers)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, txt=f"Datos de Hoja {num_hoja} ({config['nombre']})", ln=True, align="L")
    add_table_to_pdf(pdf, headers, data_for_this_factor, col_widths)
    pdf.ln(5)

# --- Main Processing Function (Refactored) ---
def process_excel_to_pdf_streamlit(uploaded_file_object):
    global user_logs
    user_logs = []

    try:
        wb = openpyxl.load_workbook(uploaded_file_object, data_only=True)
    except Exception as e:
        log_to_streamlit_and_console(f"Error al cargar el archivo Excel: {e}", "error")
        return None, None, user_logs

    pdf = FPDF(orientation='L', unit='mm', format='A3')
    pdf.set_auto_page_break(auto=True, margin=15)

    agentes_riesgo_ordenados = ["Repetitividad", "Postura", "MMC LDT", "MMC EA", "MMP", "Vibración CC", "Vibración MB"]
    resumen_riesgos_data = {} # Este diccionario se pasará por referencia y se llenará

    # --- Configuración de Hojas ---
    mapeo_pdf_hoja1 = {
        "1. ANTECEDENTES DE LA EMPRESA": {"Razón Social": (15, 'E'), "RUT Empresa": (15, 'L'), "Actividad Económica": (17, 'E'), "Código CIIU": (17, 'L'), "Dirección": (19, 'E'), "Comuna": (19, 'L'), "Nombre Representante Legal": (21, 'E'), "Organismo administrador al que está adherido": (23, 'E'), "Fecha inicio": (23, 'L')},
        "2. CENTRO DE TRABAJO O LUGAR DE TRABAJO": {"Nombre del centro de trabajo": (27, 'E'), "Dirección": (29, 'E'), "Comuna": (29, 'L'), "Nº Trabajadores Hombres": (31, 'G'), "Nº Trabajadores Mujeres": (31, 'L')},
        "3. RESPONSABLE IMPLEMENTACIÓN PROTOCOLO": {"Nombre responsable": (35, 'E'), "Cargo": (37, 'E'), "Correo electrónico": (39, 'E'), "Teléfono": (39, 'L')}
    }
    hoja2_headers_display = ["N°", "Área de trabajo", "Puesto de trabajo", "Tareas del puesto", "Descripción de la tarea", "Horario de funcionamiento", "HHEX dia", "HHEX sem", "N° trab exp hombre", "N° trab exp mujer", "Tipo contrato", "Tipo remuneracion", "Duración (min)", "Pausas", "Rotación", "Equipos - Herramientas", "Características ambientes - espacios trabajo", "Características disposición espacial puesto", "Características herramientas"]
    hoja2_col_groups_structured = [
        [(hoja2_headers_display[i], i) for i in range(5)], [(hoja2_headers_display[i], i) for i in range(5, 10)],
        [(hoja2_headers_display[i], i) for i in range(10, 14)], [(hoja2_headers_display[i], i) for i in range(14, 19)]
    ]
    hojas_factores_config = {
        "4": {"nombre": "Repetitividad", "r_col": 24, "r_filas": (14, 116), "agente_idx": 0},
        "5": {"nombre": "Postura", "r_col": 49, "r_filas": (15, 117), "agente_idx": 1},
        "6": {"nombre": "MMC LDT", "r_col": 56, "r_filas": (16, 118), "agente_idx": 2},
        "7": {"nombre": "MMC EA", "r_col": 41, "r_filas": (15, 117), "agente_idx": 3},
        "8": {"nombre": "MMP", "r_col": 41, "r_filas": (15, 117), "agente_idx": 4},
        "9": {"nombre": "Vibración CC", "r_col": 19, "r_filas": (14, 116), "agente_idx": 5},
        "10": {"nombre": "Vibración MB", "r_col": 22, "r_filas": (14, 116), "agente_idx": 6},
    }

    # --- Fase de Recolección de Datos ---
    collected_h1_data = collect_hoja1_data(wb, mapeo_pdf_hoja1)
    collected_h2_data, _ = collect_hoja2_data(wb, hoja2_col_groups_structured, resumen_riesgos_data, agentes_riesgo_ordenados) # resumen_riesgos_data se llena aquí
    
    collected_factors_data = {}
    for num_hoja_factor, config_factor in hojas_factores_config.items():
        data_for_this_factor = collect_factor_data(wb, num_hoja_factor, config_factor, resumen_riesgos_data, agentes_riesgo_ordenados) # resumen_riesgos_data se actualiza aquí
        collected_factors_data[num_hoja_factor] = data_for_this_factor
    
    # --- Fase de Escritura del PDF ---
    pdf.add_page() # Primera página
    write_hoja1_to_pdf(pdf, collected_h1_data)

    pdf.add_page() # Nueva página para el Resumen
    add_resumen_page_to_pdf(pdf, resumen_riesgos_data, agentes_riesgo_ordenados)

    pdf.add_page() # Nueva página para Hoja 2
    write_hoja2_to_pdf(pdf, collected_h2_data)

    for num_hoja_factor, config_factor in hojas_factores_config.items():
        if collected_factors_data[num_hoja_factor]: # Solo añadir página y escribir si hay datos
            pdf.add_page() # Nueva página para cada hoja de factor
            write_factor_data_to_pdf(pdf, num_hoja_factor, config_factor, collected_factors_data[num_hoja_factor])

    # --- Finalizar PDF ---
    try:
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        base_name = os.path.splitext(uploaded_file_object.name)[0]
        log_to_streamlit_and_console("PDF generado en memoria.", "info")
        return pdf_bytes, base_name, user_logs
    except Exception as e:
        log_to_streamlit_and_console(f"Error al generar bytes del PDF: {e}", "error")
        return None, None, user_logs

# --- Streamlit UI (sin cambios significativos aquí) ---
st.set_page_config(layout="wide")
st.title("📄 Convertidor Excel TMERT a PDF")
st.markdown("Cargue su archivo Excel TMERT para generar un informe PDF detallado.")

uploaded_file = st.file_uploader("📂 Cargar archivo Excel (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    st.success(f"Archivo cargado: **{uploaded_file.name}** ({uploaded_file.size / 1024:.2f} KB)")

    if st.button("🚀 Generar PDF", type="primary", use_container_width=True):
        with st.spinner("⚙️ Procesando archivo y generando PDF... Por favor espere."):
            pdf_bytes, report_base_name, logs = process_excel_to_pdf_streamlit(uploaded_file)

        st.subheader("📝 Registros del Proceso:")
        log_container = st.expander("Mostrar/Ocultar Logs", expanded=True if any("[ERROR]" in l or "[WARNING]" in l for l in logs) else False)
        with log_container:
            for log_entry in logs:
                if "[ERROR]" in log_entry:
                    st.error(log_entry)
                elif "[WARNING]" in log_entry:
                    st.warning(log_entry)
                else:
                    st.info(log_entry)
        
        if pdf_bytes and report_base_name:
            st.balloons()
            st.success("🎉 ¡PDF generado con éxito!")
            
            st.download_button(
                label="📥 Descargar PDF Generado",
                data=pdf_bytes,
                file_name=f"{report_base_name}_exportado.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        else:
            st.error("❌ No se pudo generar el PDF. Revise los logs para más detalles.")
else:
    st.info("Por favor, cargue un archivo Excel para comenzar.")

st.markdown("---")
st.markdown("Desarrollado por Mauricio Reyes Gónzalez - Especialista en Ergonomía ACHS Zona Sur (mareyesg@achs.cl), para optimizar la creación de informes TMERT.")

