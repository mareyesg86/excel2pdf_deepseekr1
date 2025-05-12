import streamlit as st
import openpyxl
# import pandas as pd # Descomenta si tu función process_excel_to_json lo necesita
import openpyxl.utils
import os
import json
from datetime import datetime
import traceback
import re # Para normalize_key
from docxtpl import DocxTemplate
from io import BytesIO # Para manejar el archivo docx en memoria para descarga

# --- Función para Normalizar Claves (la que ya tenías) ---
def normalize_key(text):
    if not isinstance(text, str):
        text = str(text)
    text = text.lower()
    replacements = {
        " ": "_", "º": "nro", ".": "", ":": "", "ñ": "n", "ó": "o", "ö": "o",
        "é": "e", "í": "i", "á": "a", "ú": "u", "ü": "u", "-": "_", "(": "", ")": "",
        "/": "_"
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    text = re.sub(r'_+', '_', text)
    return text.strip("_")

# --- Función para Procesar y Enriquecer Datos (la que ya tenías) ---
def procesar_y_enriquecer_datos(datos_crudos):
    if not datos_crudos:
        return None
    datos_procesados = datos_crudos.copy()
    if "informacion_general" in datos_procesados and "centro_trabajo" in datos_procesados["informacion_general"]:
        ct = datos_procesados["informacion_general"]["centro_trabajo"]
        try:
            hombres_ct = int(ct.get("nnro_trabajadores_hombres", 0) or 0)
            mujeres_ct = int(ct.get("nnro_trabajadores_mujeres", 0) or 0)
            ct["total_trabajadores_ct"] = hombres_ct + mujeres_ct
        except ValueError:
            ct["total_trabajadores_ct"] = "N/A (Error conversión)"
    if "puestos_trabajo_detalle" in datos_procesados:
        for puesto in datos_procesados["puestos_trabajo_detalle"]:
            try:
                hombres_exp = int(puesto.get("n°_trab_exp_hombre", 0) or 0)
                mujeres_exp = int(puesto.get("n°_trab_exp_mujer", 0) or 0)
                puesto["total_trabajadores_expuestos_puesto"] = hombres_exp + mujeres_exp
            except ValueError:
                puesto["total_trabajadores_expuestos_puesto"] = "N/A (Error conversión)"
            except TypeError:
                 puesto["total_trabajadores_expuestos_puesto"] = "N/A (Error tipo)"
    return datos_procesados

# --- Función para Generar el DOCX en memoria ---
def generar_docx_en_memoria(plantilla_bytes, contexto):
    try:
        # Cargar la plantilla desde bytes
        doc = DocxTemplate(plantilla_bytes)
        doc.render(contexto)
        
        # Guardar el documento en un buffer de bytes en memoria
        file_stream = BytesIO()
        doc.save(file_stream)
        file_stream.seek(0) # Rebobinar el stream al principio
        return file_stream
    except Exception as e:
        st.error(f"Error al generar el documento Word: {e}")
        traceback.print_exc()
        return None

# (Continuación del script app.py)

def excel_a_estructura_json(uploaded_excel_file):
    """
    Procesa un archivo Excel subido (objeto BytesIO o similar de Streamlit)
    y devuelve la estructura de datos Python (diccionario) lista para ser JSON.
    """
    if uploaded_excel_file is None:
        return None

    try:
        # openpyxl puede leer directamente desde un objeto de archivo en memoria
        wb = openpyxl.load_workbook(uploaded_excel_file, data_only=True)
    except Exception as e:
        st.error(f"Error al abrir el archivo Excel: {e}")
        traceback.print_exc()
        return None

    datos_para_json = {
        "metadata": {
            "nombre_archivo_original": uploaded_excel_file.name if hasattr(uploaded_excel_file, 'name') else "archivo_excel_cargado.xlsx",
            "fecha_procesamiento": datetime.now().isoformat()
        },
        "informacion_general": {
            "antecedentes_empresa": {},
            "centro_trabajo": {},
            "responsable_protocolo": {}
        },
        "puestos_trabajo_detalle": [],
        "resumen_global_riesgos_tabla": [] # Se llenará después
    }
    mapa_nro_puesto_a_indice_json = {}
    agentes_riesgo_ordenados = [
        "Repetitividad", "Postura", "MMC LDT", "MMC EA",
        "MMP", "Vibración MB", "Vibración CC"
    ]

    # --- Procesamiento Hoja 1 (Información General) ---
    mapeo_hoja1 = {
        "1. ANTECEDENTES DE LA EMPRESA": {
            "Razón Social": (15, 'E'), "RUT Empresa": (15, 'L'),
            "Actividad Económica": (17, 'E'), "Código CIIU": (17, 'L'),
            "Dirección": (19, 'E'), "Comuna": (19, 'L'),
            "Nombre Representante Legal": (21, 'E'),
            "Organismo administrador al que está adherido": (23, 'E'), "Fecha inicio": (23, 'L')
        },
        "2. CENTRO DE TRABAJO O LUGAR DE TRABAJO": {
            "Nombre del centro de trabajo": (27, 'E'),
            "Dirección": (29, 'E'), "Comuna": (29, 'L'),
            "Nº Trabajadores Hombres": (31, 'G'), "Nº Trabajadores Mujeres": (31, 'L')
        },
        "3. RESPONSABLE IMPLEMENTACIÓN PROTOCOLO": {
            "Nombre responsable": (35, 'E'), "Cargo": (37, 'E'),
            "Correo electrónico": (39, 'E'), "Teléfono": (39, 'L')
        }
    }
    try:
        hoja1_openpyxl = wb["1"]
        # st.write("[INFO Hoja 1] Leyendo Hoja 1.") # Puedes usar st.write para logs en Streamlit
        for seccion_titulo, campos in mapeo_hoja1.items():
            for etiqueta, (fila_excel, col_excel_char) in campos.items():
                valor_crudo_h1 = hoja1_openpyxl[f"{col_excel_char}{fila_excel}"].value
                valor_str_h1 = str(valor_crudo_h1).strip() if valor_crudo_h1 is not None and str(valor_crudo_h1).strip() != "0" else ""
                if valor_str_h1:
                    clave_json = normalize_key(etiqueta)
                    if seccion_titulo == "1. ANTECEDENTES DE LA EMPRESA":
                        datos_para_json["informacion_general"]["antecedentes_empresa"][clave_json] = valor_str_h1
                    elif seccion_titulo == "2. CENTRO DE TRABAJO O LUGAR DE TRABAJO":
                        datos_para_json["informacion_general"]["centro_trabajo"][clave_json] = valor_str_h1
                    elif seccion_titulo == "3. RESPONSABLE IMPLEMENTACIÓN PROTOCOLO":
                        datos_para_json["informacion_general"]["responsable_protocolo"][clave_json] = valor_str_h1
    except KeyError:
        st.warning("Advertencia: No se encontró la Hoja '1' en el Excel. Se omitirá esta sección.")
    except Exception as e:
        st.error(f"Error procesando Hoja '1' del Excel: {e}")
        traceback.print_exc()

    # --- Procesamiento Hoja 2 (Caracterización de Puestos) ---
    COL_NRO_H2, COL_AREA_H2, COL_PUESTO_H2, COL_TAREA_H2 = 2, 3, 4, 5
    try:
        hoja2 = wb["2"]
        hoja2_headers = [
            "N°", "Área de trabajo", "Puesto de trabajo", "Tareas del puesto",
            "Descripción de la tarea", "Horario de funcionamiento", "HHEX dia", "HHEX sem",
            "N° trab exp hombre", "N° trab exp mujer", "Tipo contrato", "Tipo remuneracion",
            "Duración (min)", "Pausas", "Rotación", "Equipos - Herramientas",
            "Características ambientes - espacios trabajo",
            "Características disposición espacial puesto", "Características herramientas"
        ]
        # st.write("[INFO Hoja 2] Leyendo Hoja 2.")
        for fila_idx in range(13, 114):
            nro_puesto_val = str(hoja2.cell(row=fila_idx, column=COL_NRO_H2).value or "").strip()
            val_a_obj = hoja2.cell(row=fila_idx, column=COL_AREA_H2).value
            val_p_obj = hoja2.cell(row=fila_idx, column=COL_PUESTO_H2).value
            val_a_str = str(val_a_obj).strip() if val_a_obj is not None else ""
            val_p_str = str(val_p_obj).strip() if val_p_obj is not None else ""

            if not (nro_puesto_val and nro_puesto_val != "0" and val_a_str and val_a_str != "0" and val_p_str and val_p_str != "0"):
                continue
            
            current_row_values = [str(hoja2.cell(row=fila_idx, column=col_idx).value or "") for col_idx in range(COL_NRO_H2, COL_NRO_H2 + len(hoja2_headers))]
            
            if any(val.strip() for val in current_row_values):
                puesto_detalle_json = {
                    normalize_key(hoja2_headers[i]): current_row_values[i]
                    for i in range(len(hoja2_headers)) if i < len(current_row_values)
                }
                puesto_detalle_json["niveles_riesgo_agentes"] = {normalize_key(agente): "AUSENTE" for agente in agentes_riesgo_ordenados}
                datos_para_json["puestos_trabajo_detalle"].append(puesto_detalle_json)
                mapa_nro_puesto_a_indice_json[nro_puesto_val] = len(datos_para_json["puestos_trabajo_detalle"]) - 1
    except KeyError:
        st.warning("Advertencia: No se encontró la Hoja '2' en el Excel. Se omitirá la caracterización de puestos.")
    except Exception as e:
        st.error(f"Error procesando Hoja '2' del Excel: {e}")
        traceback.print_exc()

    # --- Procesamiento de Hojas de Factores (4 a 10) ---
    config_hojas_factores = {
        "4": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[0]), "col_q_idx": 17, "col_x_idx": 24, "r_filas": (14, 116)},
        "5": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[1]), "col_q_idx": 31, "col_x_idx": 49, "r_filas": (17, 116)},
        "6": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[2]), "col_q_idx": 33, "col_x_idx": 56, "r_filas": (18, 118)},
        "7": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[3]), "col_q_idx": 24, "col_x_idx": 41, "r_filas": (17, 117)},
        "8": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[4]), "col_q_idx": 25, "col_x_idx": 41, "r_filas": (17, 117)},
        "9": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[5]), "col_riesgo_directo_idx": 19, "r_filas": (16, 116)}, # Vibración MB
        "10": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[6]), "col_riesgo_directo_idx": 22, "r_filas": (16, 116)}  # Vibración CC
    }
    for num_hoja_str, config in config_hojas_factores.items():
        try:
            hoja_actual = wb[num_hoja_str]
            # st.write(f"[INFO Hoja {num_hoja_str}] Leyendo Hoja {num_hoja_str} - Agente: {config['nombre_json_agente']}.")
            COL_NRO_FACTOR, COL_AREA_FACTOR, COL_PUESTO_FACTOR = 2, 3, 4
            for fila_idx in range(config["r_filas"][0], config["r_filas"][1]):
                nro_puesto_riesgo = str(hoja_actual.cell(row=fila_idx, column=COL_NRO_FACTOR).value or "").strip()
                val_a_obj = hoja_actual.cell(row=fila_idx, column=COL_AREA_FACTOR).value
                val_p_obj = hoja_actual.cell(row=fila_idx, column=COL_PUESTO_FACTOR).value
                val_a_str = str(val_a_obj).strip() if val_a_obj is not None else ""
                val_p_str = str(val_p_obj).strip() if val_p_obj is not None else ""
                if not (nro_puesto_riesgo and nro_puesto_riesgo != "0" and val_a_str and val_a_str != "0" and val_p_str and val_p_str != "0"):
                    continue
                risk_level_text = "No Determinado"
                if "col_riesgo_directo_idx" in config:
                    col_riesgo_idx = config["col_riesgo_directo_idx"]
                    valor_crudo = hoja_actual.cell(row=fila_idx, column=col_riesgo_idx).value
                    valor_str_norm = str(valor_crudo).strip().lower() if valor_crudo is not None else ""
                    if valor_str_norm == "aceptable": risk_level_text = "ACEPTABLE"
                    elif valor_str_norm == "no aceptable": risk_level_text = "CRÍTICO"
                    elif valor_str_norm: risk_level_text = str(valor_crudo).strip().upper()
                else:
                    col_q_idx = config["col_q_idx"]
                    col_x_idx = config["col_x_idx"]
                    valor_q_crudo = hoja_actual.cell(row=fila_idx, column=col_q_idx).value
                    valor_q_str = str(valor_q_crudo).strip().lower() if valor_q_crudo is not None else ""
                    if valor_q_str == "no aceptable":
                        valor_x_crudo = hoja_actual.cell(row=fila_idx, column=col_x_idx).value
                        valor_x_str = str(valor_x_crudo).strip().lower() if valor_x_crudo is not None else ""
                        if "no crítico" in valor_x_str or "intermedio" in valor_x_str: risk_level_text = "INTERMEDIO"
                        elif "crítico" in valor_x_str: risk_level_text = "CRÍTICO"
                    elif valor_q_str == "aceptable":
                        risk_level_text = "ACEPTABLE"
                if nro_puesto_riesgo in mapa_nro_puesto_a_indice_json:
                    indice_puesto = mapa_nro_puesto_a_indice_json[nro_puesto_riesgo]
                    datos_para_json["puestos_trabajo_detalle"][indice_puesto]["niveles_riesgo_agentes"][config["nombre_json_agente"]] = risk_level_text
                # else:
                    # st.warning(f"Advertencia: N° de puesto '{nro_puesto_riesgo}' de Hoja {num_hoja_str} (Agente: {config['nombre_json_agente']}) no encontrado en caracterizaciones de Hoja 2.")
        except KeyError:
            st.warning(f"Advertencia: No se encontró la Hoja '{num_hoja_str}' en el Excel. Se omitirá este factor de riesgo.")
        except Exception as e:
            st.error(f"Error procesando Hoja '{num_hoja_str}' del Excel: {e}")
            traceback.print_exc()

    # --- Crear la sección resumen_global_riesgos_tabla ---
    for puesto_detalle in datos_para_json["puestos_trabajo_detalle"]:
        item_resumen = {
            "nro": puesto_detalle.get(normalize_key("N°"), ""),
            "area": puesto_detalle.get(normalize_key("Área de trabajo"), ""),
            "puesto": puesto_detalle.get(normalize_key("Puesto de trabajo"), ""),
            "tarea": puesto_detalle.get(normalize_key("Tareas del puesto"), ""),
            "niveles_riesgo_agentes": puesto_detalle.get("niveles_riesgo_agentes", {})
        }
        datos_para_json["resumen_global_riesgos_tabla"].append(item_resumen)
    
    datos_para_json["resumen_global_riesgos_tabla"] = sorted(
        datos_para_json["resumen_global_riesgos_tabla"],
        key=lambda item: int(str(item["nro"]).split('.')[0]) if str(item["nro"]).replace('.','',1).isdigit() else float('inf')
    )
    return datos_para_json

# (Continuación del script app.py)

# --- Interfaz de Usuario y Lógica Principal de Streamlit ---
st.set_page_config(page_title="Generador de Informes TMERT", layout="wide")
st.title("Generador de Informes TMERT - Riesgo Postural (Método REBA)" - Desarrollado por Mauricio Reyes González

st.sidebar.header("1. Cargar Archivos")
uploaded_excel = st.sidebar.file_uploader("Cargar Matriz TMERT (Excel)", type=["xlsx"])
uploaded_template = st.sidebar.file_uploader("Cargar Plantilla Word (.docx)", type=["docx"])

# Variables para almacenar los datos procesados y el archivo generado
if 'datos_json_procesados' not in st.session_state:
    st.session_state.datos_json_procesados = None
if 'informe_word_bytes' not in st.session_state:
    st.session_state.informe_word_bytes = None
if 'nombre_archivo_salida' not in st.session_state:
    st.session_state.nombre_archivo_salida = "informe_tmert.docx"


if uploaded_excel and uploaded_template:
    st.sidebar.success("¡Archivos cargados!")

    if st.button("Procesar Excel y Generar Informe Filtrado (Postura INTERMEDIO)"):
        with st.spinner("Procesando Excel y generando estructura JSON..."):
            datos_crudos_json = excel_a_estructura_json(uploaded_excel)
        
        if datos_crudos_json:
            st.success("Estructura JSON generada desde Excel.")
            
            with st.spinner("Enriqueciendo datos y aplicando filtros..."):
                datos_enriquecidos = procesar_y_enriquecer_datos(datos_crudos_json)

                # --- Filtrado específico para POSTURA == 'INTERMEDIO' ---
                puestos_originales = datos_enriquecidos.get('puestos_trabajo_detalle', [])
                resumen_original = datos_enriquecidos.get('resumen_global_riesgos_tabla', [])

                puestos_filtrados = [
                    puesto for puesto in puestos_originales
                    if puesto.get('niveles_riesgo_agentes', {}).get('postura') == 'INTERMEDIO'
                ]
                numeros_puestos_filtrados = {puesto.get('n°') for puesto in puestos_filtrados}
                resumen_filtrado = [
                    resumen for resumen in resumen_original
                    if resumen.get('nro') in numeros_puestos_filtrados
                ]
                st.info(f"Filtrado aplicado: Se incluirán {len(puestos_filtrados)} puestos con riesgo de POSTURA INTERMEDIO.")
                # --- Fin Filtrado ---

                contexto_final = {
                    'metadata': datos_enriquecidos.get('metadata', {}),
                    'informacion_general': datos_enriquecidos.get('informacion_general', {}),
                    'puestos_trabajo_detalle': puestos_filtrados,
                    'resumen_global_riesgos_tabla': resumen_filtrado,
                    'fecha_actual_reporte': datetime.now().strftime("%d de %B de %Y")
                }
                st.session_state.datos_json_procesados = contexto_final # Guardar para posible inspección
            
            with st.spinner("Generando informe Word..."):
                # El archivo de plantilla se lee como bytes
                plantilla_bytes = BytesIO(uploaded_template.getvalue())
                informe_bytes = generar_docx_en_memoria(plantilla_bytes, contexto_final)
                
                if informe_bytes:
                    st.session_state.informe_word_bytes = informe_bytes
                    
                    # Definir nombre del archivo de salida
                    base_name_excel = os.path.splitext(uploaded_excel.name)[0]
                    st.session_state.nombre_archivo_salida = f"Informe_TMERT_{base_name_excel}_Postura_Intermedio.docx"
                    st.success("¡Informe Word generado y listo para descargar!")
                else:
                    st.error("No se pudo generar el informe Word.")
        else:
            st.error("No se pudo procesar el archivo Excel para generar la estructura JSON.")

# Botón de descarga (solo se muestra si hay un informe generado)
if st.session_state.informe_word_bytes:
    st.download_button(
        label="Descargar Informe Word (.docx)",
        data=st.session_state.informe_word_bytes,
        file_name=st.session_state.nombre_archivo_salida,
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )

st.markdown("---")
st.subheader("Instrucciones:")
st.markdown("""
1.  Carga tu archivo de Matriz TMERT en formato Excel (.xlsx) usando el panel de la izquierda.
2.  Carga tu plantilla de informe en formato Word (.docx) usando el panel de la izquierda.
3.  Haz clic en el botón "**Procesar Excel y Generar Informe Filtrado (Postura INTERMEDIO)**".
4.  Espera a que el proceso termine. Se generará un informe Word que incluirá solo los puestos con riesgo de **Postura INTERMEDIO**.
5.  Una vez generado, aparecerá un botón para descargar el informe.
""")

# Opcional: Mostrar el JSON procesado para depuración
if st.session_state.datos_json_procesados:
    if st.checkbox("Mostrar datos JSON procesados (para depuración)"):
        st.json(st.session_state.datos_json_procesados)
