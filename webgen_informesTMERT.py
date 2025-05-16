import streamlit as st
import openpyxl
# import pandas as pd # Descomentado si alguna función interna lo usa
import openpyxl.utils
import os
import json
from datetime import datetime, date # Añadido date para el input de fecha
import traceback
import re # Para normalize_key
from docxtpl import DocxTemplate
from io import BytesIO

# --- Función para Normalizar Claves ---
def normalize_key(text):
    if not isinstance(text, str):
        text = str(text)
    text = text.lower()
    replacements = {
        " ": "_", "º": "nro", ".": "", ":": "", "ñ": "n", "ó": "o", "ö": "o",
        "é": "e", "í": "i", "á": "a", "ú": "u", "ü": "u", "-": "_", "(": "", ")": "",
        "/": "_"
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    text = re.sub(r'_+', '_', text)
    return text.strip("_")

# --- Función para Procesar y Enriquecer Datos (Cálculos de Totales) ---
def procesar_y_enriquecer_datos(datos_crudos):
    if not datos_crudos:
        return None
    datos_procesados = datos_crudos.copy()
    if "informacion_general" in datos_procesados and "centro_trabajo" in datos_procesados["informacion_general"]:
        ct = datos_procesados["informacion_general"]["centro_trabajo"]
        try:
            hombres_ct = int(ct.get("nnro_trabajadores_hombres", 0) or 0)
            mujeres_ct = int(ct.get("nnro_trabajadores_mujeres", 0) or 0)
            ct["total_trabajadores_ct"] = hombres_ct + mujeres_ct
        except ValueError:
            ct["total_trabajadores_ct"] = "N/A (Error conversión)"
    if "puestos_trabajo_detalle" in datos_procesados:
        for puesto in datos_procesados["puestos_trabajo_detalle"]:
            try:
                hombres_exp = int(puesto.get("n°_trab_exp_hombre", 0) or 0)
                mujeres_exp = int(puesto.get("n°_trab_exp_mujer", 0) or 0)
                puesto["total_trabajadores_expuestos_puesto"] = hombres_exp + mujeres_exp
            except ValueError:
                puesto["total_trabajadores_expuestos_puesto"] = "N/A (Error conversión)"
            except TypeError:
                 puesto["total_trabajadores_expuestos_puesto"] = "N/A (Error tipo)"
    return datos_procesados

# --- Función para Generar el DOCX en memoria ---
def generar_docx_en_memoria(plantilla_bytes_io, contexto_render): # Renombrado para claridad
    try:
        doc = DocxTemplate(plantilla_bytes_io) # DocxTemplate espera un path o un file-like object
        doc.render(contexto_render)
        file_stream = BytesIO()
        doc.save(file_stream)
        file_stream.seek(0)
        return file_stream
    except Exception as e:
        st.error(f"Error al generar el documento Word: {e}")
        traceback.print_exc()
        return None

# --- Función para Procesar el Excel a la Estructura JSON ---
def excel_a_estructura_json(uploaded_excel_file):
    if uploaded_excel_file is None:
        return None
    try:
        wb = openpyxl.load_workbook(uploaded_excel_file, data_only=True)
    except Exception as e:
        st.error(f"Error al abrir el archivo Excel: {e}")
        traceback.print_exc()
        return None

    datos_para_json = {
        "metadata": {
            "nombre_archivo_original": uploaded_excel_file.name if hasattr(uploaded_excel_file, 'name') else "archivo_excel_cargado.xlsx",
            "fecha_procesamiento": datetime.now().isoformat()
        },
        "informacion_general": {
            "antecedentes_empresa": {},
            "centro_trabajo": {},
            "responsable_protocolo": {}
        },
        "puestos_trabajo_detalle": [],
        "resumen_global_riesgos_tabla": []
    }
    mapa_nro_puesto_a_indice_json = {}
    agentes_riesgo_ordenados = [
        "Repetitividad", "Postura", "MMC LDT", "MMC EA",
        "MMP", "Vibración MB", "Vibración CC"
    ]

    mapeo_hoja1 = {
        "1. ANTECEDENTES DE LA EMPRESA": {"Razón Social": (15, 'E'), "RUT Empresa": (15, 'L'), "Actividad Económica": (17, 'E'), "Código CIIU": (17, 'L'), "Dirección": (19, 'E'), "Comuna": (19, 'L'), "Nombre Representante Legal": (21, 'E'), "Organismo administrador al que está adherido": (23, 'E'), "Fecha inicio": (23, 'L')},
        "2. CENTRO DE TRABAJO O LUGAR DE TRABAJO": {"Nombre del centro de trabajo": (27, 'E'), "Dirección": (29, 'E'), "Comuna": (29, 'L'), "Nº Trabajadores Hombres": (31, 'G'), "Nº Trabajadores Mujeres": (31, 'L')},
        "3. RESPONSABLE IMPLEMENTACIÓN PROTOCOLO": {"Nombre responsable": (35, 'E'), "Cargo": (37, 'E'), "Correo electrónico": (39, 'E'), "Teléfono": (39, 'L')}
    }
    try:
        hoja1_openpyxl = wb["1"]
        for seccion_titulo, campos in mapeo_hoja1.items():
            for etiqueta, (fila_excel, col_excel_char) in campos.items():
                valor_crudo_h1 = hoja1_openpyxl[f"{col_excel_char}{fila_excel}"].value
                valor_str_h1 = str(valor_crudo_h1).strip() if valor_crudo_h1 is not None and str(valor_crudo_h1).strip() != "0" else ""
                if valor_str_h1:
                    clave_json = normalize_key(etiqueta)
                    if seccion_titulo == "1. ANTECEDENTES DE LA EMPRESA":
                        datos_para_json["informacion_general"]["antecedentes_empresa"][clave_json] = valor_str_h1
                    elif seccion_titulo == "2. CENTRO DE TRABAJO O LUGAR DE TRABAJO":
                        datos_para_json["informacion_general"]["centro_trabajo"][clave_json] = valor_str_h1
                    elif seccion_titulo == "3. RESPONSABLE IMPLEMENTACIÓN PROTOCOLO":
                        datos_para_json["informacion_general"]["responsable_protocolo"][clave_json] = valor_str_h1
    except KeyError: st.warning("Advertencia: No se encontró la Hoja '1'.")
    except Exception as e: st.error(f"Error procesando Hoja '1': {e}"); traceback.print_exc()

    COL_NRO_H2, COL_AREA_H2, COL_PUESTO_H2 = 2, 3, 4
    try:
        hoja2 = wb["2"]
        hoja2_headers = ["N°", "Área de trabajo", "Puesto de trabajo", "Tareas del puesto", "Descripción de la tarea", "Horario de funcionamiento", "HHEX dia", "HHEX sem", "N° trab exp hombre", "N° trab exp mujer", "Tipo contrato", "Tipo remuneracion", "Duración (min)", "Pausas", "Rotación", "Equipos - Herramientas", "Características ambientes - espacios trabajo", "Características disposición espacial puesto", "Características herramientas"]
        for fila_idx in range(13, 114):
            nro_puesto_val = str(hoja2.cell(row=fila_idx, column=COL_NRO_H2).value or "").strip()
            val_a_obj = hoja2.cell(row=fila_idx, column=COL_AREA_H2).value
            val_p_obj = hoja2.cell(row=fila_idx, column=COL_PUESTO_H2).value
            val_a_str = str(val_a_obj).strip() if val_a_obj is not None else ""
            val_p_str = str(val_p_obj).strip() if val_p_obj is not None else ""
            if not (nro_puesto_val and nro_puesto_val != "0" and val_a_str and val_a_str != "0" and val_p_str and val_p_str != "0"): continue
            current_row_values = [str(hoja2.cell(row=fila_idx, column=col_idx).value or "") for col_idx in range(COL_NRO_H2, COL_NRO_H2 + len(hoja2_headers))]
            if any(val.strip() for val in current_row_values):
                puesto_detalle_json = {normalize_key(hoja2_headers[i]): current_row_values[i] for i in range(len(hoja2_headers)) if i < len(current_row_values)}
                puesto_detalle_json["niveles_riesgo_agentes"] = {normalize_key(agente): "AUSENTE" for agente in agentes_riesgo_ordenados}
                datos_para_json["puestos_trabajo_detalle"].append(puesto_detalle_json)
                mapa_nro_puesto_a_indice_json[nro_puesto_val] = len(datos_para_json["puestos_trabajo_detalle"]) - 1
    except KeyError: st.warning("Advertencia: No se encontró la Hoja '2'.")
    except Exception as e: st.error(f"Error procesando Hoja '2': {e}"); traceback.print_exc()

    config_hojas_factores = {
        "4": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[0]), "col_q_idx": 17, "col_x_idx": 24, "r_filas": (14, 116)},
        "5": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[1]), "col_q_idx": 31, "col_x_idx": 49, "r_filas": (17, 116)},
        "6": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[2]), "col_q_idx": 33, "col_x_idx": 56, "r_filas": (18, 118)},
        "7": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[3]), "col_q_idx": 24, "col_x_idx": 41, "r_filas": (17, 117)},
        "8": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[4]), "col_q_idx": 25, "col_x_idx": 41, "r_filas": (17, 117)},
        "9": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[5]), "col_riesgo_directo_idx": 19, "r_filas": (16, 116)}, # Vibración MB
        "10": {"nombre_json_agente": normalize_key(agentes_riesgo_ordenados[6]), "col_riesgo_directo_idx": 22, "r_filas": (16, 116)}  # Vibración CC
    }
    for num_hoja_str, config in config_hojas_factores.items():
        try:
            hoja_actual = wb[num_hoja_str]
            COL_NRO_FACTOR, COL_AREA_FACTOR, COL_PUESTO_FACTOR = 2, 3, 4
            for fila_idx in range(config["r_filas"][0], config["r_filas"][1]):
                nro_puesto_riesgo = str(hoja_actual.cell(row=fila_idx, column=COL_NRO_FACTOR).value or "").strip()
                val_a_obj = hoja_actual.cell(row=fila_idx, column=COL_AREA_FACTOR).value
                val_p_obj = hoja_actual.cell(row=fila_idx, column=COL_PUESTO_FACTOR).value
                val_a_str = str(val_a_obj).strip() if val_a_obj is not None else ""
                val_p_str = str(val_p_obj).strip() if val_p_obj is not None else ""
                if not (nro_puesto_riesgo and nro_puesto_riesgo != "0" and val_a_str and val_a_str != "0" and val_p_str and val_p_str != "0"): continue
                risk_level_text = "No Determinado"
                if "col_riesgo_directo_idx" in config:
                    valor_crudo = hoja_actual.cell(row=fila_idx, column=config["col_riesgo_directo_idx"]).value
                    valor_str_norm = str(valor_crudo).strip().lower() if valor_crudo is not None else ""
                    if valor_str_norm == "aceptable": risk_level_text = "ACEPTABLE"
                    elif valor_str_norm == "no aceptable": risk_level_text = "CRÍTICO"
                    elif valor_str_norm: risk_level_text = str(valor_crudo).strip().upper()
                else:
                    valor_q_crudo = hoja_actual.cell(row=fila_idx, column=config["col_q_idx"]).value
                    valor_q_str = str(valor_q_crudo).strip().lower() if valor_q_crudo is not None else ""
                    if valor_q_str == "no aceptable":
                        valor_x_crudo = hoja_actual.cell(row=fila_idx, column=config["col_x_idx"]).value
                        valor_x_str = str(valor_x_crudo).strip().lower() if valor_x_crudo is not None else ""
                        if "no crítico" in valor_x_str or "intermedio" in valor_x_str: risk_level_text = "INTERMEDIO"
                        elif "crítico" in valor_x_str: risk_level_text = "CRÍTICO"
                    elif valor_q_str == "aceptable": risk_level_text = "ACEPTABLE"
                if nro_puesto_riesgo in mapa_nro_puesto_a_indice_json:
                    indice_puesto = mapa_nro_puesto_a_indice_json[nro_puesto_riesgo]
                    datos_para_json["puestos_trabajo_detalle"][indice_puesto]["niveles_riesgo_agentes"][config["nombre_json_agente"]] = risk_level_text
        except KeyError: st.warning(f"Advertencia: No se encontró la Hoja '{num_hoja_str}'.")
        except Exception as e: st.error(f"Error procesando Hoja '{num_hoja_str}': {e}"); traceback.print_exc()

    for puesto_detalle in datos_para_json["puestos_trabajo_detalle"]:
        item_resumen = {"nro": puesto_detalle.get(normalize_key("N°"), ""), "area": puesto_detalle.get(normalize_key("Área de trabajo"), ""), "puesto": puesto_detalle.get(normalize_key("Puesto de trabajo"), ""), "tarea": puesto_detalle.get(normalize_key("Tareas del puesto"), ""), "niveles_riesgo_agentes": puesto_detalle.get("niveles_riesgo_agentes", {})}
        datos_para_json["resumen_global_riesgos_tabla"].append(item_resumen)
    datos_para_json["resumen_global_riesgos_tabla"] = sorted(datos_para_json["resumen_global_riesgos_tabla"], key=lambda item: int(str(item["nro"]).split('.')[0]) if str(item["nro"]).replace('.','',1).isdigit() else float('inf'))
    return datos_para_json

# --- Interfaz de Usuario y Lógica Principal de Streamlit ---
st.set_page_config(page_title="Generador Informes TMERT", layout="wide")
st.title("Generador Dinámico de Informes TMERT 📄")

opciones_si_no = ["", "Si", "No"]
opciones_rol_empresa = ["", "Empresa principal", "Contratista", "Subcontratista", "Servicios Transitorios"]
agentes_para_filtro = ["Postura", "Repetitividad", "MMC LDT", "MMC EA", "MMP", "Vibración MB", "Vibración CC"]

col_carga, col_manual, col_accion = st.columns([2, 3, 2])

with col_carga:
    st.subheader("1. Cargar Archivo Excel 📤")
    uploaded_excel = st.file_uploader("Matriz TMERT (Excel)", type=["xlsx"], key="excel_uploader")
    # El file_uploader para la plantilla Word se elimina, ya que se cargará desde el repo

with col_manual:
    st.subheader("2. Datos Manuales (Opcional) ✍️")
    numero_informe = st.text_input("Número de Informe Técnico:", key="num_informe")
    nombre_ergonomo = st.text_input("Nombre de Ergónomo:", key="nom_ergonomo")
    rut_ergonomo = st.text_input("RUT de Ergónomo:", key="rut_ergonomo")
    correo_ergonomo = st.text_input("Correo de Ergónomo:", key="mail_ergonomo")
    fecha_visita_empresa_input = st.date_input("Fecha de Visita a Empresa:", value=None, help="Dejar en blanco si no aplica.", key="fecha_visita")
    horas_semanales_experto = st.text_input("Horas semanales de Experto Empresa:", key="hrs_experto")
    fecha_inicio_ct_input = st.date_input("Fecha Inicio CT (Contrato/Tarea):", value=None, help="Dejar en blanco si no aplica.", key="fecha_inicio_ct")
    fecha_termino_conocido_ct_input = st.date_input("Fecha Término conocido CT:", value=None, help="Dejar en blanco si no aplica.", key="fecha_termino_ct")
    fecha_termino_informe_input = st.date_input("Fecha Término (Informe):", value=None, help="Dejar en blanco si no aplica.", key="fecha_termino_informe")
    st.markdown("---")
    reglamento_hs = st.selectbox("Reglamento HS:", options=opciones_si_no, key="reg_hs")
    depto_preventivo = st.selectbox("Depto. Preventivo:", options=opciones_si_no, key="depto_prev")
    rol_empresa_ct = st.selectbox("Rol empresa en CT:", options=opciones_rol_empresa, key="rol_empresa")
    comite_paritario = st.selectbox("Comité Paritario:", options=opciones_si_no, key="comite_par")
    experto_prevencion = st.selectbox("Experto en prevención:", options=opciones_si_no, key="exp_prev")

with col_accion:
    st.subheader("3. Generar Informe ⚙️")
    agente_seleccionado_filtro = st.selectbox(
        "Filtrar por Factor de Riesgo (Nivel INTERMEDIO):",
        options=agentes_para_filtro, index=0, key="agente_filtro"
    )
    
    # --- MAPEO DE AGENTE A NOMBRE DE ARCHIVO DE PLANTILLA ---
    # !!! IMPORTANTE: AJUSTA ESTE MAPEO SEGÚN TUS NOMBRES DE ARCHIVO Y AGENTES !!!
    MAPEO_AGENTE_A_PLANTILLA = {
        normalize_key("Repetitividad"): "plantillas/1_ART.docx", # CONFIRMA SI 1_ART.docx ES PARA REPETITIVIDAD
        normalize_key("Postura"): "plantillas/2_REBA.docx",       # CONFIRMA SI 2_REBA.docx ES PARA POSTURA
        normalize_key("MMC LDT"): "plantillas/3_MAC.docx",
        normalize_key("MMC EA"): "plantillas/4_RAPP.docx",
        # Añade aquí los mapeos para las otras plantillas cuando las tengas
        # normalize_key("MMP"): "plantillas/plantilla_mmp.docx",
        # normalize_key("Vibración MB"): "plantillas/plantilla_vibracion_mb.docx",
        # normalize_key("Vibración CC"): "plantillas/plantilla_vibracion_cc.docx"
    }
    
    if st.button(f"🚀 Procesar y Generar Informe", key="generate_button"):
        if uploaded_excel: # Solo necesitamos el Excel ahora
            with st.spinner("⚙️ Procesando Excel..."):
                datos_crudos_json = excel_a_estructura_json(uploaded_excel)
            
            if datos_crudos_json:
                st.success("✅ Estructura JSON generada.")
                with st.spinner("🔍 Aplicando cálculos y filtros..."):
                    datos_enriquecidos = procesar_y_enriquecer_datos(datos_crudos_json)
                    clave_agente_filtro_json = normalize_key(agente_seleccionado_filtro)
                    puestos_originales = datos_enriquecidos.get('puestos_trabajo_detalle', [])
                    resumen_original = datos_enriquecidos.get('resumen_global_riesgos_tabla', [])
                    puestos_filtrados = [puesto for puesto in puestos_originales if puesto.get('niveles_riesgo_agentes', {}).get(clave_agente_filtro_json) == 'INTERMEDIO']
                    numeros_puestos_filtrados = {puesto.get(normalize_key('N°')) for puesto in puestos_filtrados}
                    resumen_filtrado = [resumen for resumen in resumen_original if resumen.get('nro') in numeros_puestos_filtrados]
                    st.info(f"📊 Filtro aplicado: Se incluirán {len(puestos_filtrados)} puestos con riesgo de {agente_seleccionado_filtro} INTERMEDIO.")

                    contexto_final = {
                        'metadata': datos_enriquecidos.get('metadata', {}),
                        'informacion_general': datos_enriquecidos.get('informacion_general', {}),
                        'puestos_trabajo_detalle': puestos_filtrados,
                        'resumen_global_riesgos_tabla': resumen_filtrado,
                        'fecha_actual_reporte': datetime.now().strftime("%d de %B de %Y"),
                        'numero_informe_tecnico': numero_informe,
                        'nombre_ergonomo': nombre_ergonomo,
                        'rut_ergonomo': rut_ergonomo,
                        'correo_ergonomo': correo_ergonomo,
                        'fecha_visita_empresa': fecha_visita_empresa_input.strftime("%d-%m-%Y") if fecha_visita_empresa_input else "",
                        'horas_semanales_experto_empresa': horas_semanales_experto,
                        'fecha_inicio_ct': fecha_inicio_ct_input.strftime("%d-%m-%Y") if fecha_inicio_ct_input else "",
                        'fecha_termino_conocido_ct': fecha_termino_conocido_ct_input.strftime("%d-%m-%Y") if fecha_termino_conocido_ct_input else "",
                        'fecha_termino_informe': fecha_termino_informe_input.strftime("%d-%m-%Y") if fecha_termino_informe_input else "",
                        'reglamento_hs': reglamento_hs,
                        'depto_preventivo': depto_preventivo,
                        'rol_empresa_en_ct': rol_empresa_ct,
                        'comite_paritario': comite_paritario,
                        'experto_en_prevencion': experto_prevencion
                    }
                
                # --- Cargar la plantilla Word seleccionada dinámicamente ---
                nombre_agente_norm_para_plantilla = normalize_key(agente_seleccionado_filtro)
                ruta_plantilla_en_repo = MAPEO_AGENTE_A_PLANTILLA.get(nombre_agente_norm_para_plantilla)

                if not ruta_plantilla_en_repo:
                    st.error(f"❌ No se encontró una plantilla mapeada para el agente: '{agente_seleccionado_filtro}'. Verifica el diccionario 'MAPEO_AGENTE_A_PLANTILLA' en el código y los archivos en la carpeta 'plantillas'.")
                else:
                    with st.spinner(f"📄 Cargando plantilla '{ruta_plantilla_en_repo}' y generando informe Word..."):
                        try:
                            with open(ruta_plantilla_en_repo, "rb") as f_template:
                                plantilla_bytes_seleccionada = BytesIO(f_template.read())
                            
                            informe_bytes = generar_docx_en_memoria(plantilla_bytes_seleccionada, contexto_final)
                            
                            if informe_bytes:
                                base_name_excel = os.path.splitext(uploaded_excel.name)[0]
                                num_informe_para_nombre = numero_informe.strip()
                                
                                if num_informe_para_nombre:
                                    prefijo_nombre = f"Informe_TMERT_Nro_{num_informe_para_nombre.replace('/', '_').replace(' ', '_')}"
                                else:
                                    nombre_excel_limpio = base_name_excel.replace(' ', '_')
                                    prefijo_nombre = f"Informe_TMERT_{nombre_excel_limpio}"
                                
                                nombre_archivo_salida = f"{prefijo_nombre}_{agente_seleccionado_filtro.replace(' ','_')}_Intermedio.docx"
                                
                                st.download_button(
                                    label=f"📥 Descargar Informe ({agente_seleccionado_filtro} INTERMEDIO)",
                                    data=informe_bytes,
                                    file_name=nombre_archivo_salida,
                                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                    key="download_button_final_mapeado"
                                )
                                st.success("🎉 ¡Informe Word generado!")
                            else:
                                st.error("❌ No se pudo generar el informe Word con la plantilla seleccionada.")
                        except FileNotFoundError:
                            st.error(f"❌ Error crítico: No se encontró la plantilla '{ruta_plantilla_en_repo}' en el repositorio. Verifica que el archivo exista en la carpeta 'plantillas' y que el nombre en el mapeo sea exacto.")
                        except Exception as e_template_proc:
                            st.error(f"❌ Error al cargar o procesar la plantilla específica: {e_template_proc}")
                            traceback.print_exc()
            else:
                st.error("❌ No se pudo procesar el archivo Excel.")
        else:
            st.warning("⚠️ Por favor, carga el archivo Excel.")

# Opcional: Mostrar el JSON procesado para depuración
# if 'contexto_final' in locals():
#     if st.checkbox("Mostrar datos JSON completos para el informe (depuración)", key="show_json_checkbox"):
#         st.json(contexto_final)
